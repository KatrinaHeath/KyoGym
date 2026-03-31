"""
Script para sincronizar la base de datos SQLite con Excel en OneDrive
usando Microsoft Graph API
"""
import sqlite3
from pathlib import Path
from datetime import datetime
import json
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl no está instalado. Ejecuta: pip install openpyxl")
    exit(1)

try:
    import msal
    import requests
except ImportError:
    print("Error: msal no está instalado. Ejecuta: pip install msal requests")
    exit(1)

from utils.constants import DB_PATH


# ==================== CONFIGURACIÓN ====================
# Archivo de configuración para credenciales
CONFIG_FILE = Path(__file__).parent / "onedrive_config.json"

# Configuración por defecto (se puede sobrescribir con el archivo de configuración)
DEFAULT_CONFIG = {
    "client_id": "TU_CLIENT_ID",
    "client_secret": "TU_CLIENT_SECRET",
    "tenant_id": "TU_TENANT_ID",
    "authority": "https://login.microsoftonline.com/TU_TENANT_ID",
    "scope": ["https://graph.microsoft.com/.default"],
    "onedrive_folder": "/",  # Ruta en OneDrive donde se guardará el archivo
    "excel_filename": "gimnasio.xlsx"
}


class OneDriveSync:
    """Clase para sincronizar la base de datos con OneDrive"""
    
    def __init__(self, config_path=None):
        """Inicializa el sincronizador con la configuración"""
        self.config = self._load_config(config_path)
        self.access_token = None
        
    def _load_config(self, config_path=None):
        """Carga la configuración desde archivo o usa valores por defecto"""
        if config_path is None:
            config_path = CONFIG_FILE
            
        if config_path.exists():
            print(f"📁 Cargando configuración desde {config_path}")
            with open(config_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        else:
            print(f"⚠️  Archivo de configuración no encontrado: {config_path}")
            print(f"📝 Creando archivo de configuración por defecto...")
            self._create_default_config(config_path)
            return DEFAULT_CONFIG
    
    def _create_default_config(self, config_path):
        """Crea un archivo de configuración por defecto"""
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(DEFAULT_CONFIG, f, indent=4, ensure_ascii=False)
        print(f"✅ Archivo de configuración creado: {config_path}")
        print(f"⚠️  IMPORTANTE: Edita este archivo con tus credenciales de Azure AD")
    
    def authenticate(self):
        """Autentica con Microsoft Graph API usando flujo de credenciales de cliente"""
        print("\n🔐 Autenticando con Microsoft Graph API...")
        
        # Validar configuración
        if self.config.get("client_id") == "TU_CLIENT_ID":
            raise ValueError(
                "❌ Error: Debes configurar tus credenciales en onedrive_config.json\n"
                "Visita: https://portal.azure.com para crear una aplicación Azure AD"
            )
        
        try:
            # Crear aplicación MSAL
            app = msal.ConfidentialClientApplication(
                self.config["client_id"],
                authority=self.config["authority"],
                client_credential=self.config["client_secret"]
            )
            
            # Obtener token de acceso
            result = app.acquire_token_for_client(scopes=self.config["scope"])
            
            if "access_token" in result:
                self.access_token = result["access_token"]
                print("✅ Autenticación exitosa")
                return True
            else:
                error_msg = result.get("error_description", result.get("error"))
                raise Exception(f"Error de autenticación: {error_msg}")
                
        except Exception as e:
            print(f"❌ Error al autenticar: {str(e)}")
            raise
    
    def read_database(self):
        """Lee los datos de las tablas de la base de datos"""
        print(f"\n📊 Leyendo datos de la base de datos: {DB_PATH}")
        
        if not DB_PATH.exists():
            raise FileNotFoundError(f"Base de datos no encontrada: {DB_PATH}")
        
        conn = sqlite3.connect(str(DB_PATH))
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        data = {}
        
        # Leer clientes
        cursor.execute("SELECT * FROM clientes ORDER BY id")
        data['clientes'] = [dict(row) for row in cursor.fetchall()]
        print(f"  ✓ Clientes: {len(data['clientes'])} registros")
        
        # Leer membresías
        cursor.execute("""
            SELECT m.*, c.nombre as cliente_nombre 
            FROM membresias m
            LEFT JOIN clientes c ON m.cliente_id = c.id
            ORDER BY m.id
        """)
        data['membresias'] = [dict(row) for row in cursor.fetchall()]
        print(f"  ✓ Membresías: {len(data['membresias'])} registros")
        
        # Leer pagos
        cursor.execute("""
            SELECT p.*, c.nombre as cliente_nombre 
            FROM pagos p
            LEFT JOIN clientes c ON p.cliente_id = c.id
            ORDER BY p.id
        """)
        data['pagos'] = [dict(row) for row in cursor.fetchall()]
        print(f"  ✓ Pagos: {len(data['pagos'])} registros")
        
        conn.close()
        return data
    
    def create_excel(self, data, output_path):
        """Crea un archivo Excel con los datos de la base de datos"""
        print(f"\n📝 Creando archivo Excel: {output_path}")
        
        wb = Workbook()
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Hoja 1: Clientes
        ws_clientes = wb.active
        ws_clientes.title = "Clientes"
        
        if data['clientes']:
            headers = list(data['clientes'][0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws_clientes.cell(1, col, header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            for row_idx, cliente in enumerate(data['clientes'], 2):
                for col_idx, header in enumerate(headers, 1):
                    ws_clientes.cell(row_idx, col_idx, cliente[header])
            
            # Ajustar ancho de columnas
            for col in range(1, len(headers) + 1):
                ws_clientes.column_dimensions[get_column_letter(col)].width = 15
        
        print(f"  ✓ Hoja 'Clientes' creada")
        
        # Hoja 2: Membresías
        ws_membresias = wb.create_sheet("Membresías")
        
        if data['membresias']:
            headers = list(data['membresias'][0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws_membresias.cell(1, col, header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            for row_idx, membresia in enumerate(data['membresias'], 2):
                for col_idx, header in enumerate(headers, 1):
                    ws_membresias.cell(row_idx, col_idx, membresia[header])
            
            # Ajustar ancho de columnas
            for col in range(1, len(headers) + 1):
                ws_membresias.column_dimensions[get_column_letter(col)].width = 15
        
        print(f"  ✓ Hoja 'Membresías' creada")
        
        # Hoja 3: Pagos
        ws_pagos = wb.create_sheet("Pagos")
        
        if data['pagos']:
            headers = list(data['pagos'][0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws_pagos.cell(1, col, header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            for row_idx, pago in enumerate(data['pagos'], 2):
                for col_idx, header in enumerate(headers, 1):
                    ws_pagos.cell(row_idx, col_idx, pago[header])
            
            # Ajustar ancho de columnas
            for col in range(1, len(headers) + 1):
                ws_pagos.column_dimensions[get_column_letter(col)].width = 15
        
        print(f"  ✓ Hoja 'Pagos' creada")
        
        # Guardar archivo
        wb.save(output_path)
        print(f"✅ Archivo Excel creado exitosamente")
        
        return output_path
    
    def upload_to_onedrive(self, file_path):
        """Sube el archivo Excel a OneDrive"""
        print(f"\n☁️  Subiendo archivo a OneDrive...")
        
        if not self.access_token:
            raise Exception("No hay token de acceso. Ejecuta authenticate() primero")
        
        filename = self.config.get("excel_filename", "gimnasio.xlsx")
        onedrive_folder = self.config.get("onedrive_folder", "/")
        
        # Construir la ruta completa en OneDrive
        if onedrive_folder == "/":
            upload_path = f"/me/drive/root:/{filename}:/content"
        else:
            upload_path = f"/me/drive/root:{onedrive_folder}/{filename}:/content"
        
        url = f"https://graph.microsoft.com/v1.0{upload_path}"
        
        headers = {
            "Authorization": f"Bearer {self.access_token}",
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }
        
        # Leer el archivo
        with open(file_path, 'rb') as f:
            file_data = f.read()
        
        # Subir el archivo
        try:
            response = requests.put(url, headers=headers, data=file_data)
            response.raise_for_status()
            
            result = response.json()
            print(f"✅ Archivo subido exitosamente a OneDrive")
            print(f"   📁 Nombre: {result.get('name')}")
            print(f"   🔗 ID: {result.get('id')}")
            print(f"   📏 Tamaño: {result.get('size')} bytes")
            
            # Si hay una URL web, mostrarla
            if 'webUrl' in result:
                print(f"   🌐 URL: {result.get('webUrl')}")
            
            return result
            
        except requests.exceptions.HTTPError as e:
            error_detail = e.response.json() if e.response.text else str(e)
            print(f"❌ Error al subir archivo: {error_detail}")
            raise
        except Exception as e:
            print(f"❌ Error inesperado: {str(e)}")
            raise
    
    def sync(self):
        """Ejecuta el proceso completo de sincronización"""
        print("=" * 60)
        print("🔄 SINCRONIZACIÓN GIMNASIO.DB → ONEDRIVE")
        print("=" * 60)
        
        try:
            # 1. Autenticar
            self.authenticate()
            
            # 2. Leer datos de la base de datos
            data = self.read_database()
            
            # 3. Crear archivo Excel temporal
            temp_excel = Path(__file__).parent / "temp_gimnasio.xlsx"
            self.create_excel(data, temp_excel)
            
            # 4. Subir a OneDrive
            self.upload_to_onedrive(temp_excel)
            
            # 5. Limpiar archivo temporal
            if temp_excel.exists():
                temp_excel.unlink()
                print(f"\n🗑️  Archivo temporal eliminado")
            
            print("\n" + "=" * 60)
            print("✅ SINCRONIZACIÓN COMPLETADA EXITOSAMENTE")
            print(f"⏰ Fecha y hora: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            print("=" * 60)
            
            return True
            
        except Exception as e:
            print(f"\n❌ Error durante la sincronización: {str(e)}")
            return False


def main():
    """Función principal"""
    try:
        # Crear instancia del sincronizador
        syncer = OneDriveSync()
        
        # Ejecutar sincronización
        syncer.sync()
        
    except KeyboardInterrupt:
        print("\n\n⚠️  Sincronización cancelada por el usuario")
    except Exception as e:
        print(f"\n❌ Error fatal: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
